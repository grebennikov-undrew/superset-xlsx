# Licensed to the Apache Software Foundation (ASF) under one
# or more contributor license agreements.  See the NOTICE file
# distributed with this work for additional information
# regarding copyright ownership.  The ASF licenses this file
# to you under the Apache License, Version 2.0 (the
# "License"); you may not use this file except in compliance
# with the License.  You may obtain a copy of the License at
#
#   http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing,
# software distributed under the License is distributed on an
# "AS IS" BASIS, WITHOUT WARRANTIES OR CONDITIONS OF ANY
# KIND, either express or implied.  See the License for the
# specific language governing permissions and limitations
# under the License.
import io
import re
from typing import Any, Dict
from superset import (
    app as superset_app,
)
from openpyxl.utils.cell import get_column_letter
from openpyxl.styles import Border, PatternFill, Alignment, Color, Font
from openpyxl.styles.borders import Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles.differential import DifferentialStyle
import pandas as pd
import openpyxl
import logging

logger = logging.getLogger(__name__)
config = superset_app.config

# Formats mapping (d3 -> xlsx)
def get_xlsx_num_type(d3_type):
    xlsx_type = ''
    if d3_type[-1]=='%':
      # Percentage format
      # EXAMPLE: ,.2% -> 0,00%
      xlsx_type = '0.' + ('0'*int(d3_type[-2])) + '%'
      return xlsx_type
    if d3_type[-1]=='f':
      # Float format (only with delimeters)
      # EXAMPLE: ,.2f -> # ##0,00
      xlsx_type = '# ##0.' + ('0'*int(d3_type[-2]))
      return xlsx_type
    if d3_type=='d':
      # Rounded integers
      # EXAMPLE: d -> 0
      xlsx_type = '0'
      return xlsx_type
    #return 'General'
    return xlsx_type

def get_xlsx_date_type(d3_type):
    xlsx_type = ''
    # If DTTM
    if 'd' in d3_type and 'm' in d3_type and 'Y' in d3_type and 'H' in d3_type and 'M' in d3_type and 'S' in d3_type:
        xlsx_type = 'DD.MM.YYYY HH:MM:SS'
    # If only date
    elif 'd' in d3_type and 'm' in d3_type and 'Y' in d3_type:
        xlsx_type ='DD.MM.YYYY'
    # If time only
    elif 'H' in d3_type and 'M' in d3_type and 'S' in d3_type:
        xlsx_type = 'HH:MM:SS'
    # else:
    #     xlsx_type = 'General'
    return xlsx_type


def get_xlsx_formats(superset_formats):
    #xlsx_formats = dict.fromkeys(superset_formats.keys())
    xlsx_formats = {}
    # Map formats from SS to XLSX. Each column
    for column, formats in superset_formats.items():
        column_format = {}
        # Each SS format for one column
        for format,value in formats.items():
            # Map numeric formats (include percentage)
            if format == 'd3NumberFormat':
                column_format['datatype'] = get_xlsx_num_type(value)
            elif format == 'd3TimeFormat':
                column_format['datatype'] = get_xlsx_date_type(value)
            #else:
            #    column_format['datatype'] = 'General'
        xlsx_formats[column] = column_format
    return xlsx_formats

def get_xlsx_conditional(superset_conditions):
    xlsx_conditions = []
    operators_mapping = {'>': 'greaterThan', 
                         '<': 'lessThan',
                         '≥': 'greaterThanOrEqual',
                         '≤': 'lessThanOrEqual',
                         '=': 'equal',
                         '≠': 'notEqual',
                         '< x <': 'between'}
    for condition in  superset_conditions:
        column_name = condition['column']
        color = condition['colorScheme'][1:]
        fill = PatternFill(fill_type='solid', start_color=color, end_color=color)
        operator = condition['operator']
        if (operator == '>' or operator == '<' or operator == '≥' or operator == '≤' or operator == '=' or operator == '≠'):
            value = condition['targetValue']
            xlsx_conditions.append( {column_name: {'type': 'cell',
                                            'operator': operators_mapping[operator],
                                            'formula':    [value],
                                            'fill':   fill}})
        elif (operator == '< x <'):
            minimum = float(condition['targetValueLeft']) + 0.0001
            maximum = float(condition['targetValueRight']) - 0.0001
            #xlsx_conditions[column_name] = {'type':     'cell',
            xlsx_conditions.append( {column_name: {'type': 'cell',
                                            'operator': operators_mapping[operator],
                                            'formula':    [minimum,maximum],
                                            'fill':   fill}})
    return xlsx_conditions

# Returns width of xlsx columns.
# Rules:  
#     - Column width is 75 percentiles of char count
#     - Min width is 13
#     - Max width is 56
#     - Width of date columns = 14
#     - If header width is greater than 75 percentiles by no more than 30%, column will have width of header
def get_column_width(df):
    width = {}
    for column in df.columns:
        header_width = len(column)
        min_width = 12
        max_width = 42
        if df.dtypes[column] == 'datetime64[ns]':
            data_width = 12
        else:
            data_width = df[column].map(str).map(len).quantile(0.75)*0.75

        if data_width < min_width:
            width[column] = min_width
        elif data_width > max_width:
            width[column] = max_width
        elif data_width < header_width and data_width * 1.3 > header_width:
            width[column] = header_width + 2
        else:
            width[column] = data_width + 2
    return(width)


def df_to_excel(df: pd.DataFrame, form_data: Dict[str, Any], **kwargs: Any) -> Any:
    logger.info("# form_data is: %s", form_data)
    output = io.BytesIO()
    logger.info("# Default export date format is %s", config["EXCEL_DATE_FORMAT"])
    with pd.ExcelWriter(
        output, 
        engine="openpyxl",
        datetime_format = config["EXCEL_DATE_FORMAT"],
        date_format = config["EXCEL_DATE_FORMAT"]
    ) as writer:
        # Count of rows and columns in df
        LAST_ROW = len(df.index) + 1
        COLUMN_COUNT = len(df.columns)
        print('FORM_DATA IS: ', form_data)
        # Get formatting from form_data
        superset_formats = form_data.get('column_config')
        superset_conditions = form_data.get('conditional_formatting')
        # Dict of xlsx formats and conditional formats
        xlsx_formats = get_xlsx_formats(superset_formats)
        xlsx_conditions = get_xlsx_conditional(superset_conditions)
        df.to_excel(
            writer, 
            **kwargs,
            index = False)
        workbook  = writer.book
        ws = writer.sheets['Sheet1']

        ### DATA STYLING ###

        # Implementing borders, width and word-wrap           
        thin_border = Border(left=Side(style='thin'), 
                      right=Side(style='thin'), 
                      top=Side(style='thin'), 
                      bottom=Side(style='thin'))
        
        # Column width generation
        col_width = get_column_width(df.replace(to_replace=r"<a href=(.*)>(.*)</a>", value=r"\2", regex=True))

        # Link pattern 
        link_pattern = re.compile(r"<a href=\"(.*)\">(.*)</a>")

        # FOR EACH CELL
        for column in df.columns:
            # in openpyxls column "A" has index "1"
            col_num = df.columns.get_loc(column) + 1

            # Implementing column width
            ws.column_dimensions[get_column_letter(col_num)].width = col_width[column]
            for row in ws.iter_rows(min_row=2, max_row=LAST_ROW, min_col=col_num, max_col=col_num):    
                for cell in row:

                    # Implementing datatypes
                    if column in xlsx_formats:
                        col_format = xlsx_formats[column]
                        col_datetype = col_format['datatype']
                        cell.number_format = col_datetype
                    # else:
                    #     cell.number_format = 'General'

                    # Implementing font, borders and word-wrap
                    cell.border = thin_border
                    cell.alignment = Alignment(wrapText=True,
                                               vertical='top')
                    cell.font = Font(name = 'Tahoma', size = 8)

                    # Inserting hyperlinks
                    if link_pattern.match(str(cell.value)):
                        link = link_pattern.findall(cell.value)[0][0]
                        label = link_pattern.findall(cell.value)[0][1]
                        if label.isdigit():
                            cell.hyperlink = link
                            cell.value = float(label)
                            cell.font = Font(name = 'Tahoma', 
                                            size = 8, 
                                            underline='single',
                                            color='000000FF')

                   
            ws.column_dimensions[get_column_letter(col_num)].auto_size = True

        # Implementing condtional formatting
        for condition in xlsx_conditions:
            for column, description in condition.items():
                # in openpyxls column "A" has index "1"
                col_num = df.columns.get_loc(column) + 1
                # xlsx letter range
                col_letter = get_column_letter(col_num)
                condition_area = f"{col_letter}2:{col_letter}{LAST_ROW}"
                # adding conditions based on formula
                ws.conditional_formatting.add(condition_area,
                    CellIsRule(operator=description['operator'], 
                               formula=description['formula'], 
                               fill=description['fill'],
                               stopIfTrue=True,))

        # Implementing autofilters
        ws.auto_filter.ref = ws.dimensions

        ### HEADER STYLING ###
        ws.row_dimensions[1].height = 28
        header_format = {'bold': False,
                          'text_wrap': True,
                          'align': 'center',
                          'valign': 'center',
                          'fg_color': 'CFE0F1',
                          'border': 1}
        for column in df.columns:
            # in openpyxls column "A" has index "1"
            col_num = df.columns.get_loc(column) + 1
            col_letter = get_column_letter(col_num)
            # for A1, B1 ...
            cell = ws[f"{col_letter}1"]
            cell.alignment = Alignment(horizontal=header_format['align'],
                                      vertical=header_format['valign'],
                                      wrapText=True)
            cell.border = thin_border
            cell.font = Font(name = 'Tahoma',
                             size = 8,
                             bold = header_format['bold'])
            cell.fill = PatternFill('solid', fgColor=header_format['fg_color'])

    return output.getvalue()